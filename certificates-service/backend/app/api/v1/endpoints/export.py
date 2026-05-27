import io
from typing import Optional
from datetime import date

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.certificate import Certificate, CertificateStatus

router = APIRouter()


@router.get("/excel")
async def export_to_excel(
    q: Optional[str] = None,
    product_type: Optional[str] = None,
    material: Optional[str] = None,
    gost: Optional[str] = None,
    manufacturer: Optional[str] = None,
    status: Optional[CertificateStatus] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Certificate)

    if q and q.strip():
        q_clean = q.strip()
        query = query.where(or_(
            Certificate.normalized_product_name.ilike(f"%{q_clean}%"),
            Certificate.material.ilike(f"%{q_clean}%"),
            Certificate.gost.ilike(f"%{q_clean}%"),
            Certificate.manufacturer.ilike(f"%{q_clean}%"),
            Certificate.certificate_number.ilike(f"%{q_clean}%"),
        ))
    if product_type:
        query = query.where(Certificate.product_type.ilike(f"%{product_type}%"))
    if material:
        query = query.where(Certificate.material.ilike(f"%{material}%"))
    if gost:
        query = query.where(Certificate.gost.ilike(f"%{gost}%"))
    if manufacturer:
        query = query.where(Certificate.manufacturer.ilike(f"%{manufacturer}%"))
    if status:
        query = query.where(Certificate.status == status)
    if date_from:
        query = query.where(Certificate.certificate_date >= date_from)
    if date_to:
        query = query.where(Certificate.certificate_date <= date_to)

    query = query.order_by(Certificate.certificate_date.desc().nullslast())
    result = await db.execute(query)
    certs = result.scalars().all()

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сертификаты"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Дата сертификата",
        "Номер сертификата",
        "Товар",
        "Тип продукции",
        "Размер",
        "Марка стали",
        "ГОСТ / ТУ",
        "Производитель",
        "Поставщик",
        "Номер партии/плавки",
        "Статус",
        "Файл",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[1].height = 25

    for row_num, cert in enumerate(certs, 2):
        row_data = [
            cert.certificate_date.strftime("%d.%m.%Y") if cert.certificate_date else "",
            cert.certificate_number or "",
            cert.normalized_product_name or cert.product_name or "",
            cert.product_type or "",
            cert.dimensions or "",
            cert.material or "",
            cert.gost or "",
            cert.manufacturer or "",
            cert.supplier or "",
            cert.batch_number or cert.heat_number or "",
            cert.status.value if cert.status else "",
            cert.original_filename or "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        if row_num % 2 == 0:
            alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = alt_fill

    # Auto-fit columns
    col_widths = [15, 20, 45, 20, 15, 15, 18, 30, 25, 20, 12, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Output
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=certificates.xlsx"},
    )
